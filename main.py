import requests
import json
from openpyxl import Workbook
import pandas

response = requests.post("http://rnd-orel.site/test/")

response.encoding = 'utf-8-sig'
resp = response.json()

keys = list(resp.keys())

def write(data, filename):
    data = json.dumps(data)
    data = json.loads(data)
    with open(filename, 'w', encoding='utf-8') as file:
        json.dump(data, file, indent=8)


guid = {keys[-1]: resp[keys[-1]]}
write (guid, 'last_element.json')


#_________________________________________

book = Workbook()
sheet = book.active

n = 0
i = 1
j = 2
m = 0
k = 2
sheet.cell(row=1, column=1).value = 'keys'
sheet.cell(row=1, column=2).value = 'value'
for resp1 in keys:
    sheet.cell(row=j, column=i).value = keys[n]

    if type(resp[keys[m]]) == list:
        t = 0
        list = resp[keys[m]]
        while t < len(resp[keys[m]]):
             sheet.cell(row=k, column=i + 1 + t).value = list[t]
             t += 1

    elif type(resp[keys[m]]) == dict:
        dict1 = resp[keys[m]]
        w = 0

        for dictant , amogus in dict1.items():
            sheet.cell(row=k + w, column=i + 1).value = dictant
            sheet.cell(row=k + w, column=i + 2).value = amogus
            w += 1

    else:
        sheet.cell(row=k, column=i+1).value = resp[keys[m]]

    m += 1
    k += 1
    j += 1
    n += 1





book.save('all.xlsx')
book.close()